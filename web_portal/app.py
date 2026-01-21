from flask import Flask, render_template_string, request, redirect, url_for, session, Response, jsonify
import os
import cv2
import json
import threading
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys
from datetime import datetime
from config import CAMERA_INDEX

# Add parent directory to path to import local modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import EmployeeDatabase
from config import (
    USERNAME, PASSWORD, USER_USERNAME, USER_PASSWORD,
    ATTENDANCE_FOLDER, PHOTOS_FOLDER, EMPLOYEES_DB_FILE, DATA_FOLDER
)
from models import FaceDetector, FaceRecognizer
from attendance_service import VideoRegistrationService
from camera import CameraManager

# Global detectors for web portal
detector = None
recognizer = None

def get_models():
    global detector, recognizer
    if detector is None:
        detector = FaceDetector()
    if recognizer is None:
        recognizer = FaceRecognizer()
    return detector, recognizer

# Import local modules from this folder
try:
    from web_portal.templates import (
        LOGIN_TEMPLATE, EMPLOYEES_TEMPLATE, EMPLOYEE_FORM_TEMPLATE,
        MONTHLY_TEMPLATE, DAILY_TEMPLATE,
        EMPLOYEE_DETAIL_TEMPLATE, MONTHLY_DETAIL_TEMPLATE
    )
    from web_portal.utils import AttendanceReader
except ImportError:
    from templates import (
        LOGIN_TEMPLATE, EMPLOYEES_TEMPLATE, EMPLOYEE_FORM_TEMPLATE,
        MONTHLY_TEMPLATE, DAILY_TEMPLATE,
        EMPLOYEE_DETAIL_TEMPLATE, MONTHLY_DETAIL_TEMPLATE
    )
    from utils import AttendanceReader

# Tắt log truy cập của Flask/Werkzeug để terminal sạch hơn
import logging
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

app = Flask(__name__)
app.secret_key = 'super-secret-key-123'

# Global state for background video processing (Upload mode)
processing_status = {
    'active': False,
    'done': False,
    'error': None,
    'current': 0,
    'total': 0,
    'percent': 0
}

def admin_required(f):
    def wrap(*args, **kwargs):
        if 'role' not in session or session['role'] != 'admin':
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrap.__name__ = f.__name__
    return wrap

def login_required(f):
    def wrap(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrap.__name__ = f.__name__
    return wrap

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = request.form['username']
        pwd = request.form['password']
        
        # Check Admin
        if user == USERNAME and pwd == PASSWORD:
            session['logged_in'] = True
            session['role'] = 'admin'
            return redirect(url_for('monthly_dashboard'))
        
        # Check Employee View Only Account
        if user == USER_USERNAME and pwd == USER_PASSWORD:
            session['logged_in'] = True
            session['role'] = 'user'
            return redirect(url_for('monthly_dashboard'))

        return render_template_string(LOGIN_TEMPLATE, error="Sai tài khoản hoặc mật khẩu!")
    return render_template_string(LOGIN_TEMPLATE)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return redirect(url_for('monthly_dashboard'))

@app.route('/monthly')
@login_required
def monthly_dashboard():
    reader = AttendanceReader(ATTENDANCE_FOLDER, EMPLOYEES_DB_FILE)
    data = reader.get_monthly_summary(31)
    return render_template_string(MONTHLY_TEMPLATE, data=data)

@app.route('/daily')
@login_required
def daily_dashboard():
    date_str = request.args.get('date', datetime.now().strftime("%Y-%m-%d"))
    reader = AttendanceReader(ATTENDANCE_FOLDER, EMPLOYEES_DB_FILE)
    
    available_dates = reader.get_available_dates(31)
    if not available_dates and date_str == datetime.now().strftime("%Y-%m-%d"):
        available_dates = [date_str]
    
    records = reader.get_records_by_date(date_str)
    
    # Calculate stats
    total = len(records)
    on_time = sum(1 for r in records.values() if r["status"] == "Đúng giờ")
    issues = sum(1 for r in records.values() if r["status"] not in ["Đúng giờ", "Nghỉ"])
    percent = (on_time / total * 100) if total > 0 else 0
    
    stats = {
        "total": total,
        "on_time": on_time,
        "issues": issues,
        "on_time_percent": percent
    }
    
    return render_template_string(DAILY_TEMPLATE, 
                                 employees=records, 
                                 date=date_str, 
                                 available_dates=available_dates,
                                 stats=stats)

@app.route('/monthly-detail')
@login_required
def monthly_detail():
    days = int(request.args.get('days', 31))
    reader = AttendanceReader(ATTENDANCE_FOLDER, EMPLOYEES_DB_FILE)
    dates = reader.get_available_dates(days)
    dates.sort() # Show chronological order in detail view
    
    employees = reader.get_employees_info()
    employee_data = {}
    
    for emp_id in employees:
        employee_data[emp_id] = {}
        for date in dates:
            day_records = reader.get_records_by_date(date)
            if emp_id in day_records:
                employee_data[emp_id][date] = day_records[emp_id]

    return render_template_string(MONTHLY_DETAIL_TEMPLATE, 
                                 employees=employees, 
                                 dates=dates, 
                                 employee_data=employee_data,
                                 days=days)

@app.route('/employee/<emp_id>')
@login_required
def employee_detail(emp_id):
    reader = AttendanceReader(ATTENDANCE_FOLDER, EMPLOYEES_DB_FILE)
    employees = reader.get_employees_info()
    
    if emp_id not in employees:
        return "Nhân viên không tồn tại", 404
        
    info = employees[emp_id]
    history = reader.get_employee_history(emp_id, 31)
    
    return render_template_string(EMPLOYEE_DETAIL_TEMPLATE,
                                 emp_id=emp_id,
                                 emp_name=info['name'],
                                 emp_dept=info.get('department', 'N/A'),
                                 history=history)

@app.route('/api/update-status', methods=['POST'])
@admin_required
def update_status():
    data = request.json
    emp_id = data.get('emp_id')
    date_str = data.get('date')
    new_status = data.get('status')
    
    if not all([emp_id, date_str, new_status]):
        return jsonify({"status": "error", "message": "Thiếu dữ liệu"}), 400
        
    # Ensure directory exists
    if not os.path.exists(ATTENDANCE_FOLDER):
        os.makedirs(ATTENDANCE_FOLDER)
        
    manual_path = os.path.join(ATTENDANCE_FOLDER, f"manual_status_{date_str}.json")
    
    # Load existing
    overrides = {}
    if os.path.exists(manual_path):
        try:
            with open(manual_path, 'r', encoding='utf-8') as f:
                overrides = json.load(f)
        except:
            pass
            
    # Update
    overrides[emp_id] = new_status
    
    # Save
    with open(manual_path, 'w', encoding='utf-8') as f:
        json.dump(overrides, f, indent=4, ensure_ascii=False)
        
    return jsonify({"status": "success", "message": "Đã cập nhật trạng thái"})

@app.route('/employees')
@admin_required
def list_employees():
    db = EmployeeDatabase()
    employees = db.get_all_employees()
    # Add count of photos for each employee
    for eid in employees:
        # Sử dụng _sanitize_folder_name để tìm đúng thư mục (vì tên có thể có dấu)
        safe_name = db._sanitize_folder_name(employees[eid]['name'])
        emp_dir = os.path.join(PHOTOS_FOLDER, f"{eid}_{safe_name}")
        
        if os.path.exists(emp_dir):
            employees[eid]['num_photos'] = len([f for f in os.listdir(emp_dir) if f.lower().endswith(('.jpg', '.png', '.jpeg'))])
        else:
            # Thử tìm các thư mục có prefix là eid_ (phòng trường hợp tên thư mục bị lệch)
            found = False
            if os.path.exists(PHOTOS_FOLDER):
                for folder in os.listdir(PHOTOS_FOLDER):
                    if folder.startswith(f"{eid}_"):
                        p = os.path.join(PHOTOS_FOLDER, folder)
                        employees[eid]['num_photos'] = len([f for f in os.listdir(p) if f.lower().endswith(('.jpg', '.png', '.jpeg'))])
                        found = True
                        break
            if not found:
                employees[eid]['num_photos'] = 0
            
    return render_template_string(EMPLOYEES_TEMPLATE, employees=employees)

@app.route('/employees/add', methods=['GET', 'POST'])
@admin_required
def add_employee():
    if request.method == 'POST':
        emp_id = request.form['emp_id']
        name = request.form['name']
        dept = request.form.get('department', '')
        video_file = request.files.get('video')
        
        db = EmployeeDatabase()
        if db.exists(emp_id):
            return jsonify({'status': 'error', 'message': 'Mã nhân viên đã tồn tại!'})
            
        if not video_file:
            return jsonify({'status': 'error', 'message': 'Vui lòng tải lên video!'})
            
        # Lưu file video tạm thời
        temp_video_path = os.path.join(DATA_FOLDER, f"temp_{emp_id}_{video_file.filename}")
        video_file.save(temp_video_path)
            
        # Reset progress
        global processing_status
        processing_status = {
            'active': True,
            'done': False,
            'error': None,
            'current': 0,
            'total': 0,
            'percent': 0
        }
        
        # Chạy xử lý trong thread riêng
        det, rec = get_models()
        threading.Thread(
            target=process_uploaded_video, 
            args=(temp_video_path, emp_id, name, dept, det, rec),
            daemon=True
        ).start()
        
        return jsonify({'status': 'success'})
        
    return render_template_string(EMPLOYEE_FORM_TEMPLATE, action='add', employee=None)

def process_uploaded_video(video_path, emp_id, name, dept, det, rec):
    """Hàm xử lý video upload chạy ngầm."""
    global processing_status
    try:
        cap = cv2.VideoCapture(video_path)
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        processing_status['total'] = total_frames
        
        captured_frames = []
        frame_idx = 0
        
        # Thiết lập frame skip để lấy đều các góc mặt (lấy khoảng 60 frame tiềm năng)
        # Nếu video ngắn, lấy nhiều hơn, nếu dài lấy ít hơn
        skip = max(1, total_frames // 60)
        
        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break
                
            if frame_idx % skip == 0:
                # Detect face
                face = det.detect(frame)
                if face:
                    captured_frames.append(frame.copy())
            
            frame_idx += 1
            processing_status['current'] = frame_idx
            processing_status['percent'] = int((frame_idx / total_frames) * 100)
            
        cap.release()
        
        if len(captured_frames) < 3:
            processing_status['error'] = "Không tìm thấy đủ khuôn mặt rõ trong video!"
            processing_status['active'] = False
            return
            
        # Đăng ký với database thông qua service
        db = EmployeeDatabase()
        service = VideoRegistrationService(det, rec, db)
        success, message = service.register_from_video_frames(emp_id, name, dept, captured_frames)
        
        if success:
            processing_status['done'] = True
        else:
            processing_status['error'] = message
            
    except Exception as e:
        processing_status['error'] = str(e)
    finally:
        processing_status['active'] = False
        if os.path.exists(video_path):
            try: os.remove(video_path)
            except: pass

@app.route('/employees/edit/<emp_id>', methods=['GET', 'POST'])
@admin_required
def edit_employee(emp_id):
    db = EmployeeDatabase()
    if not db.exists(emp_id):
        return redirect(url_for('list_employees'))
        
    if request.method == 'POST':
        name = request.form['name']
        dept = request.form.get('department', '')
        # Sử dụng update_employee thay vì add_employee_info_only giúp cập nhật thông tin hiện có
        success, message = db.update_employee(emp_id, name, dept)
        if success:
            return redirect(url_for('list_employees'))
        else:
            employee = db.get_employee(emp_id).copy()
            employee['emp_id'] = emp_id
            return render_template_string(EMPLOYEE_FORM_TEMPLATE, action='edit', employee=employee, error=message)
        
    employee = db.get_employee(emp_id).copy()
    employee['emp_id'] = emp_id
    return render_template_string(EMPLOYEE_FORM_TEMPLATE, action='edit', employee=employee)

@app.route('/employees/delete/<emp_id>', methods=['POST'])
@admin_required
def delete_employee(emp_id):
    db = EmployeeDatabase()
    db.delete_employee(emp_id)
    return redirect(url_for('list_employees'))

@app.route('/api/video-feed')
@admin_required
def video_feed():
    """Video feed đơn giản - chỉ để quan sát camera."""
    def generate():
        # Sử dụng CameraManager với Threading để giảm lag RTSP
        with CameraManager(CAMERA_INDEX) as camera:
            if not camera.is_opened:
                return

            while True:
                success, frame = camera.read_frame()
                if not success or frame is None:
                    time.sleep(0.01)
                    continue
                
                ret, buffer = cv2.imencode('.jpg', frame)
                frame_bytes = buffer.tobytes()
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
        
    return Response(generate(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/process-progress')
@admin_required
def process_progress():
    """API trả về tiến độ xử lý video upload."""
    return jsonify(processing_status)

@app.route('/export-excel')
@login_required
def export_excel():
    days = int(request.args.get('days', 7))
    reader = AttendanceReader(ATTENDANCE_FOLDER, EMPLOYEES_DB_FILE)
    dates = reader.get_available_dates(days)
    dates.sort() # Sắp xếp theo trình tự thời gian
    
    employees = reader.get_employees_info()
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Diem_Danh_{days}_Ngay"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Header row 1: Thông tin NV và Các ngày
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(1, 1).value = "STT"
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.cell(1, 2).value = "Mã NV"
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    ws.cell(1, 3).value = "Họ Tên"
    ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
    ws.cell(1, 4).value = "Phòng Ban"
    
    col_idx = 5
    for d in dates:
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+2)
        cell = ws.cell(1, col_idx)
        cell.value = d
        cell.alignment = center_align
        cell.font = header_font
        cell.fill = header_fill
        
        # Sub-header: Vào, Ra, Trạng thái
        ws.cell(2, col_idx).value = "Vào"
        ws.cell(2, col_idx+1).value = "Ra"
        ws.cell(2, col_idx+2).value = "Trạng thái"
        col_idx += 3
        
    # Formatting headers
    for r in [1, 2]:
        for c in range(1, col_idx):
            cell = ws.cell(r, c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

    # Data rows
    row_idx = 3
    for emp_id, info in employees.items():
        ws.cell(row_idx, 1).value = row_idx - 2
        ws.cell(row_idx, 2).value = emp_id
        ws.cell(row_idx, 3).value = info['name']
        ws.cell(row_idx, 4).value = info.get('department', 'N/A')
        
        c_idx = 5
        for d in dates:
            day_records = reader.get_records_by_date(d)
            if emp_id in day_records:
                rec = day_records[emp_id]
                ws.cell(row_idx, c_idx).value = rec['check_in'] or "--"
                ws.cell(row_idx, c_idx+1).value = rec['check_out'] or "--"
                status_cell = ws.cell(row_idx, c_idx+2)
                status_cell.value = rec['status']
                
                # Color status
                if rec['status'] == 'Đúng giờ':
                    status_cell.font = Font(color="008000") # Green
                elif rec['status'] != 'Nghỉ':
                    status_cell.font = Font(color="FF0000") # Red
                else:
                    status_cell.font = Font(color="808080") # Gray
            else:
                ws.cell(row_idx, c_idx).value = "--"
                ws.cell(row_idx, c_idx+1).value = "--"
                ws.cell(row_idx, c_idx+2).value = "Nghỉ"
                ws.cell(row_idx, c_idx+2).font = Font(color="808080")
            
            # Auto-align data
            for i in range(3):
                ws.cell(row_idx, c_idx+i).alignment = center_align
                ws.cell(row_idx, c_idx+i).border = border
                
            c_idx += 3
        
        # Border and align for static cols
        for i in range(1, 5):
            ws.cell(row_idx, i).border = border
            ws.cell(row_idx, i).alignment = center_align
            
        row_idx += 1
        
    # Column width adjustments
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    
    output_filename = f"Bao_Cao_Diem_Danh_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    temp_path = os.path.join(ATTENDANCE_FOLDER, output_filename)
    wb.save(temp_path)
    
    # Function to yield file and then delete
    def stream_and_remove():
        with open(temp_path, 'rb') as f:
            yield from f
        try:
            os.remove(temp_path)
        except:
            pass
            
    return Response(stream_and_remove(), 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={"Content-Disposition": f"attachment; filename={output_filename}"})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
